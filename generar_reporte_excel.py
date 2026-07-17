#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generar_reporte_excel.py
=========================
Recolecta los resultados (.npy) generados por las corridas estadísticas del
módulo Prey-Predator (carpeta `Results/<N>/prey_predator_metrics_report.npy`)
y los organiza en un único archivo Excel (.xlsx), fácil de leer.

Cómo funciona
-------------
1. Recorre `Results/<N>/` y carga `prey_predator_metrics_report.npy`.
   Cada archivo es un arreglo numpy de forma (R, 7): una fila por réplica,
   con las columnas (en este orden, tal como las guarda
   `raoi_simulator/prey_predator.py`):

       completion_time, success_fraction, cohesion_mean, n_captured,
       inter_capture_interval, cohesion_at_capture, swarm_split_ratio

2. Lee `Notes/*.txt` (las transcripciones de terminal que el usuario fue
   guardando) y extrae, por número de Carpeta, los parámetros con los que
   se corrió esa simulación: Predators, Preys, radios y modo de asignación
   (Emergent/Focused). Esto sirve para identificar qué corrida es cada
   Carpeta dentro del Excel (el .npy por sí solo no trae esos datos).

3. Construye un Excel con 3 hojas:
   - "Léeme"        : explicación de qué contiene cada hoja/columna.
   - "Resumen"       : una fila por Carpeta, con sus parámetros de entrada
                        y el promedio de cada métrica (calculado con
                        fórmulas de Excel =AVERAGEIF, no hardcodeado).
   - "Datos_Detalle" : una fila por réplica individual (todas las Carpetas),
                        con sus parámetros y sus 7 métricas.

Uso
---
    python generar_reporte_excel.py

Genera `Reporte_Resultados_PreyPredator.xlsx` en la raíz del repositorio.
Requiere: numpy, openpyxl  (pip install numpy openpyxl)
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------------------------- #
# Configuración
# --------------------------------------------------------------------------- #

BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR / "Results"
NOTES_DIR = BASE_DIR / "Notes"
OUTPUT_PATH = BASE_DIR / "Reporte_Resultados_PreyPredator.xlsx"
NPY_FILENAME = "prey_predator_metrics_report.npy"

# Orden real de columnas dentro de cada .npy (ver
# raoi_simulator/prey_predator.py, función run_statistical / línea
# "metrics_report[r] = [...]").
METRIC_COLUMNS = [
    ("completion_time", "Tiempo de captura (iter)", "0.00"),
    ("success_fraction", "Fracción de éxito", "0.000"),
    ("cohesion_mean", "Cohesión media (m)", "0.0000"),
    ("n_captured", "Presas capturadas", "0"),
    ("inter_capture_interval", "Intervalo entre capturas (iter)", "0.00"),
    ("cohesion_at_capture", "Cohesión al capturar (m)", "0.0000"),
    ("swarm_split_ratio", "Tasa de división del enjambre", "0.0000"),
]

# Colores / estilos
COLOR_HEADER = "1F4E78"
COLOR_HEADER_TXT = "FFFFFF"
COLOR_EA = "DDEBF7"   # azul claro -> Emergent
COLOR_FA = "FCE4D6"   # naranja claro -> Focused
FONT_NAME = "Arial"

THIN = Side(style="thin", color="B7B7B7")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------- #
# 1. Parametros por Carpeta, extraídos de Notes/*.txt
# --------------------------------------------------------------------------- #

NOTE_BLOCK_RE = re.compile(
    r"Carpeta:\s*(\d+).*?"
    r"Predators:\s*(\d+).*?"
    r"Preys:\s*(\d+).*?"
    r"Repulsion radius \(m\):\s*([\d.]+).*?"
    r"Orientation radius \(m\):\s*([\d.]+).*?"
    r"Attraction radius \(m\):\s*([\d.]+).*?"
    r"Influence radius \(m\):\s*([\d.]+).*?"
    r"Prey repulsion radius[^:]*:\s*([\d.]+).*?"
    r"Allocation mode\s*:\s*(Emergent \(EA\)|Focused \(FA\))",
    re.DOTALL,
)


def parse_notes(notes_dir: Path) -> dict[int, dict]:
    """Lee todos los .txt de Notes y arma {carpeta: {parametros...}}."""
    params: dict[int, dict] = {}
    if not notes_dir.is_dir():
        return params

    for txt_path in sorted(notes_dir.glob("*.txt")):
        text = txt_path.read_text(encoding="utf-8", errors="replace")
        for m in NOTE_BLOCK_RE.finditer(text):
            carpeta = int(m.group(1))
            params[carpeta] = {
                "predators": int(m.group(2)),
                "preys": int(m.group(3)),
                "rep_radius": float(m.group(4)),
                "ori_radius": float(m.group(5)),
                "att_radius": float(m.group(6)),
                "inf_radius": float(m.group(7)),
                "prey_rep_radius": float(m.group(8)),
                "allocation": m.group(9),
                "source_note": txt_path.name,
            }
    return params


# --------------------------------------------------------------------------- #
# 2. Carga de los .npy en Results/<N>/
# --------------------------------------------------------------------------- #

def load_results(results_dir: Path) -> dict[int, np.ndarray]:
    """Devuelve {carpeta: array (R, 7)} para cada subcarpeta con .npy válido."""
    data: dict[int, np.ndarray] = {}
    if not results_dir.is_dir():
        print(f"[!] No existe la carpeta de resultados: {results_dir}")
        return data

    for sub in sorted(results_dir.iterdir(), key=lambda p: (len(p.name), p.name)):
        if not sub.is_dir():
            continue
        npy_path = sub / NPY_FILENAME
        if not npy_path.exists():
            print(f"[!] {sub.name}: no se encontró {NPY_FILENAME}, se omite.")
            continue
        try:
            carpeta = int(sub.name)
        except ValueError:
            print(f"[!] Carpeta '{sub.name}' no es numérica, se omite.")
            continue

        arr = np.load(npy_path, allow_pickle=True)
        arr = np.atleast_2d(arr)
        if arr.shape[1] != len(METRIC_COLUMNS):
            print(
                f"[!] Carpeta {carpeta}: se esperaban {len(METRIC_COLUMNS)} "
                f"columnas y se encontraron {arr.shape[1]}. Se omite."
            )
            continue
        data[carpeta] = arr

    return data


# --------------------------------------------------------------------------- #
# 3. Construcción del Excel
# --------------------------------------------------------------------------- #

def style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT_NAME, bold=True, color=COLOR_HEADER_TXT, size=11)
        cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_ALL


def autosize_columns(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def build_readme_sheet(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Léeme"
    ws.sheet_view.showGridLines = False

    title_font = Font(name=FONT_NAME, bold=True, size=16, color=COLOR_HEADER)
    section_font = Font(name=FONT_NAME, bold=True, size=12, color=COLOR_HEADER)
    body_font = Font(name=FONT_NAME, size=11)

    ws["B2"] = "Reporte de Resultados — Tarea Prey-Predator (RAOI Swarm Simulator)"
    ws["B2"].font = title_font

    ws["B4"] = (
        "Este archivo se generó automáticamente a partir de los .npy guardados en "
        "Results/<carpeta>/prey_predator_metrics_report.npy de este repositorio, "
        "ejecutando el script generar_reporte_excel.py."
    )
    ws["B4"].font = body_font
    ws["B4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("B4:H4")
    ws.row_dimensions[4].height = 32

    rows = [
        ("", ""),
        ("Contenido de las hojas", ""),
        ("Resumen", "Una fila por Carpeta (cada Carpeta = una configuración de simulación "
                     "corrida con 'Statistical run'). Incluye los parámetros de entrada "
                     "(Predators, Preys, modo de asignación, radios), el promedio de cada "
                     "métrica sobre todas sus réplicas, su desviación estándar, y una columna "
                     "combinada 'Media ± DE' lista para reportar o copiar/pegar."),
        ("Datos_Detalle", "Una fila por réplica individual (todas las Carpetas juntas). "
                           "Permite filtrar/dinamizar los datos crudos en Excel (tablas dinámicas, etc.)."),
        ("", ""),
        ("Significado de las métricas", ""),
        ("Tiempo de captura (iter)", "completion_time: iteraciones de simulación hasta capturar todas las presas asignadas."),
        ("Fracción de éxito", "success_fraction: proporción de presas que sí fueron capturadas (1.0 = el 100%)."),
        ("Cohesión media (m)", "cohesion_mean: dispersión promedio del enjambre de depredadores durante toda la corrida (metros)."),
        ("Presas capturadas", "n_captured: número de presas efectivamente capturadas en esa réplica."),
        ("Intervalo entre capturas (iter)", "inter_capture_interval: iteraciones promedio transcurridas entre una captura y la siguiente."),
        ("Cohesión al capturar (m)", "cohesion_at_capture: cohesión del enjambre justo en el instante de cada captura (metros)."),
        ("Tasa de división del enjambre", "swarm_split_ratio: fracción de iteraciones en que el enjambre se dividió para perseguir presas distintas (0 = nunca se dividió)."),
        ("", ""),
        ("Parámetros de entrada", ""),
        ("Predators / Preys", "Número de depredadores y presas en esa configuración."),
        ("Modo de asignación", "Emergent (EA): asignación de presas surge del comportamiento del enjambre. "
                                "Focused (FA): asignación dirigida/forzada."),
        ("Radios (m)", "Repulsión, orientación, atracción, influencia y repulsión de presa: parámetros del modelo de comportamiento (zonas del enjambre)."),
        ("", ""),
        ("Fuente de los parámetros", "Predators, Preys y Allocation mode se extrajeron automáticamente de las "
                                       "transcripciones de terminal en la carpeta Notes/*.txt (columna 'Nota origen' "
                                       "en la hoja Resumen indica de qué .txt salió cada fila)."),
    ]

    r = 6
    for label, desc in rows:
        if desc == "" and label != "":
            ws.cell(row=r, column=2, value=label).font = section_font
        elif label == "" and desc == "":
            pass
        else:
            ws.cell(row=r, column=2, value=label).font = Font(name=FONT_NAME, bold=True, size=11)
            cell = ws.cell(row=r, column=3, value=desc)
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 28
        r += 1

    autosize_columns(ws, {1: 2, 2: 26, 3: 14, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14})


def build_detail_sheet(wb: Workbook, results: dict[int, np.ndarray],
                        params: dict[int, dict]) -> dict[int, tuple[int, int]]:
    """Crea la hoja Datos_Detalle. Devuelve {carpeta: (primera_fila, ultima_fila)}
    para que la hoja Resumen pueda referenciar los rangos con fórmulas."""
    ws = wb.create_sheet("Datos_Detalle")

    headers = (
        ["Carpeta", "Réplica", "Predators", "Preys", "Modo de asignación"]
        + [label for _, label, _ in METRIC_COLUMNS]
    )
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    carpeta_ranges: dict[int, tuple[int, int]] = {}
    row_idx = 2
    for carpeta in sorted(results.keys()):
        arr = results[carpeta]
        meta = params.get(carpeta, {})
        first_row = row_idx
        for replica_i, row_values in enumerate(arr, start=1):
            allocation = meta.get("allocation", "—")
            ws.cell(row=row_idx, column=1, value=carpeta)
            ws.cell(row=row_idx, column=2, value=replica_i)
            ws.cell(row=row_idx, column=3, value=meta.get("predators", "—"))
            ws.cell(row=row_idx, column=4, value=meta.get("preys", "—"))
            ws.cell(row=row_idx, column=5, value=allocation)

            fill = None
            if allocation.startswith("Emergent"):
                fill = PatternFill("solid", fgColor=COLOR_EA)
            elif allocation.startswith("Focused"):
                fill = PatternFill("solid", fgColor=COLOR_FA)

            for col_offset, (key, _, num_fmt) in enumerate(METRIC_COLUMNS):
                cell = ws.cell(row=row_idx, column=6 + col_offset, value=float(row_values[col_offset]))
                cell.number_format = num_fmt
                cell.border = BORDER_ALL
                if fill:
                    cell.fill = fill

            for col in range(1, 6):
                ws.cell(row=row_idx, column=col).border = BORDER_ALL
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")
                if fill:
                    ws.cell(row=row_idx, column=col).fill = fill

            row_idx += 1
        carpeta_ranges[carpeta] = (first_row, row_idx - 1)

    widths = {1: 10, 2: 10, 3: 11, 4: 9, 5: 20}
    for i in range(len(METRIC_COLUMNS)):
        widths[6 + i] = 24
    autosize_columns(ws, widths)

    ws.auto_filter.ref = f"A1:{get_column_letter(5 + len(METRIC_COLUMNS))}{row_idx - 1}"

    return carpeta_ranges


def build_summary_sheet(wb: Workbook, results: dict[int, np.ndarray], params: dict[int, dict],
                         carpeta_ranges: dict[int, tuple[int, int]]) -> None:
    ws = wb.create_sheet("Resumen")

    n_metrics = len(METRIC_COLUMNS)

    # Bloque de 3 columnas por métrica: Promedio | Desv. estándar | Media ± DE
    metric_headers: list[str] = []
    for _, label, _ in METRIC_COLUMNS:
        metric_headers.extend([
            "Promedio " + label,
            "Desv. estándar " + label,
            label + " (Media ± DE)",
        ])

    headers = (
        ["Carpeta", "Predators", "Preys", "Modo de asignación", "Réplicas",
         "Radio repulsión (m)", "Radio orientación (m)", "Radio atracción (m)",
         "Radio influencia (m)", "Radio repulsión presa (m)"]
        + metric_headers
        + ["Nota origen"]
    )
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    FIRST_METRIC_COL = 11  # columna donde arranca el primer bloque de métricas
    COLS_PER_METRIC = 3
    note_col = FIRST_METRIC_COL + n_metrics * COLS_PER_METRIC
    last_col = note_col

    detail_sheet = "Datos_Detalle"
    row_idx = 2
    for carpeta in sorted(results.keys()):
        meta = params.get(carpeta, {})
        allocation = meta.get("allocation", "—")
        first_row, last_row = carpeta_ranges[carpeta]
        n_replicas = last_row - first_row + 1

        ws.cell(row=row_idx, column=1, value=carpeta)
        ws.cell(row=row_idx, column=2, value=meta.get("predators", "—"))
        ws.cell(row=row_idx, column=3, value=meta.get("preys", "—"))
        ws.cell(row=row_idx, column=4, value=allocation)
        ws.cell(row=row_idx, column=5, value=n_replicas)
        ws.cell(row=row_idx, column=6, value=meta.get("rep_radius", "—"))
        ws.cell(row=row_idx, column=7, value=meta.get("ori_radius", "—"))
        ws.cell(row=row_idx, column=8, value=meta.get("att_radius", "—"))
        ws.cell(row=row_idx, column=9, value=meta.get("inf_radius", "—"))
        ws.cell(row=row_idx, column=10, value=meta.get("prey_rep_radius", "—"))

        for col_offset, (_, _, num_fmt) in enumerate(METRIC_COLUMNS):
            data_col_letter = get_column_letter(6 + col_offset)
            mean_col = FIRST_METRIC_COL + col_offset * COLS_PER_METRIC
            std_col = mean_col + 1
            combo_col = mean_col + 2

            # Como cada Carpeta ocupa un bloque contiguo de filas en
            # Datos_Detalle (first_row..last_row), se puede usar el rango
            # directo en lugar de una fórmula condicional tipo *IF.
            mean_formula = (
                f"=AVERAGEIF('{detail_sheet}'!$A${first_row}:$A${last_row},"
                f"A{row_idx},'{detail_sheet}'!{data_col_letter}${first_row}:{data_col_letter}${last_row})"
            )
            mean_cell = ws.cell(row=row_idx, column=mean_col, value=mean_formula)
            mean_cell.number_format = num_fmt
            mean_cell.border = BORDER_ALL

            if n_replicas > 1:
                std_formula = (
                    f"=STDEV('{detail_sheet}'!{data_col_letter}${first_row}:{data_col_letter}${last_row})"
                )
            else:
                # STDEV.S de una sola muestra no está definida (Excel devuele #DIV/0!)
                std_formula = 0
            std_cell = ws.cell(row=row_idx, column=std_col, value=std_formula)
            std_cell.number_format = num_fmt
            std_cell.border = BORDER_ALL

            mean_ref = f"{get_column_letter(mean_col)}{row_idx}"
            std_ref = f"{get_column_letter(std_col)}{row_idx}"
            combo_formula = (
                f'=TEXT({mean_ref},"{num_fmt}")&" ± "&TEXT({std_ref},"{num_fmt}")'
            )
            combo_cell = ws.cell(row=row_idx, column=combo_col, value=combo_formula)
            combo_cell.border = BORDER_ALL
            combo_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=note_col, value=meta.get("source_note", "—"))

        fill = None
        if allocation.startswith("Emergent"):
            fill = PatternFill("solid", fgColor=COLOR_EA)
        elif allocation.startswith("Focused"):
            fill = PatternFill("solid", fgColor=COLOR_FA)
        if fill:
            for col in range(1, last_col + 1):
                ws.cell(row=row_idx, column=col).fill = fill

        for col in range(1, last_col + 1):
            ws.cell(row=row_idx, column=col).border = BORDER_ALL
            if col != note_col:
                ws.cell(row=row_idx, column=col).alignment = Alignment(horizontal="center")

        row_idx += 1

    last_data_row = row_idx - 1
    widths = {1: 10, 2: 11, 3: 9, 4: 20, 5: 10, 6: 16, 7: 17, 8: 16, 9: 16, 10: 20}
    for i in range(n_metrics):
        base = FIRST_METRIC_COL + i * COLS_PER_METRIC
        widths[base] = 20        # Promedio
        widths[base + 1] = 20    # Desv. estándar
        widths[base + 2] = 22    # Media ± DE
    widths[note_col] = 14
    autosize_columns(ws, widths)

    ws.auto_filter.ref = f"A1:{get_column_letter(last_col)}{last_data_row}"


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #

def main() -> None:
    print(f"Buscando resultados en: {RESULTS_DIR}")
    results = load_results(RESULTS_DIR)
    if not results:
        print("[X] No se encontraron archivos .npy válidos. No se generó el Excel.")
        sys.exit(1)

    print(f"  -> {len(results)} carpeta(s) de resultados cargada(s): {sorted(results.keys())}")

    print(f"Leyendo parámetros desde: {NOTES_DIR}")
    params = parse_notes(NOTES_DIR)
    print(f"  -> Parámetros encontrados para {len(params)} carpeta(s).")

    missing = sorted(set(results.keys()) - set(params.keys()))
    if missing:
        print(f"[!] No se hallaron parámetros en Notes/ para las carpetas: {missing} "
              f"(quedarán con '—' en Predators/Preys/Modo).")

    wb = Workbook()
    build_readme_sheet(wb)
    carpeta_ranges = build_detail_sheet(wb, results, params)
    build_summary_sheet(wb, results, params, carpeta_ranges)

    wb.save(OUTPUT_PATH)
    print(f"\n[OK] Excel generado en: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
